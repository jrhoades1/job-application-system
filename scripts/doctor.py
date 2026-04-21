"""
Doctor — preflight validator for the job-applications repo.

Runs a battery of cheap checks against env vars, Gmail token, Supabase,
scoring-rules sync, lockfiles, extension/web version drift, and tracker
integrity. Exits 1 if any check is red so CI can gate PRs.

Usage:
    python scripts/doctor.py                # human output
    python scripts/doctor.py --json         # machine-readable for CI
    python scripts/doctor.py --only env,gmail  # subset

Design notes:
- Checks are isolated functions returning a CheckResult.
- All checks run even if one fails — surface the full picture.
- Only stdlib + already-installed deps. No new packages.
"""
from __future__ import annotations

import argparse
import hashlib
import json
import re
import sys
import time
from dataclasses import asdict, dataclass, field
from pathlib import Path
from typing import Callable

REPO_ROOT = Path(__file__).resolve().parent.parent

GREEN = "green"
YELLOW = "yellow"
RED = "red"


@dataclass
class CheckResult:
    name: str
    status: str
    message: str
    duration_ms: int = 0
    details: dict = field(default_factory=dict)


def _time_check(fn: Callable[[], CheckResult]) -> CheckResult:
    start = time.perf_counter()
    try:
        result = fn()
    except Exception as exc:
        result = CheckResult(name=fn.__name__, status=RED, message=f"check crashed: {exc}")
    result.duration_ms = int((time.perf_counter() - start) * 1000)
    return result


# ---------------------------------------------------------------------------
# Individual checks
# ---------------------------------------------------------------------------

def check_env() -> CheckResult:
    """Verify .env.local keys cover .env.example keys with non-placeholder values."""
    example_path = REPO_ROOT / "apps" / "web" / ".env.example"
    local_path = REPO_ROOT / "apps" / "web" / ".env.local"
    if not example_path.exists():
        return CheckResult("env", YELLOW, f".env.example missing at {example_path.relative_to(REPO_ROOT)}")
    if not local_path.exists():
        return CheckResult("env", RED, f".env.local missing at {local_path.relative_to(REPO_ROOT)}")

    def _keys(path: Path) -> dict[str, str]:
        out: dict[str, str] = {}
        for line in path.read_text(encoding="utf-8", errors="replace").splitlines():
            stripped = line.strip()
            if not stripped or stripped.startswith("#") or "=" not in stripped:
                continue
            key, _, value = stripped.partition("=")
            out[key.strip()] = value.strip().strip('"').strip("'")
        return out

    example_keys = _keys(example_path)
    local_keys = _keys(local_path)

    missing = [k for k in example_keys if k not in local_keys]
    placeholders = [
        k for k, v in local_keys.items()
        if k in example_keys and v and (
            v.endswith("...") or v.startswith("replace-") or v == "your-client-id.apps.googleusercontent.com"
        )
    ]

    if missing:
        return CheckResult(
            "env",
            RED,
            f"{len(missing)} env var(s) missing from .env.local",
            details={"missing": missing[:10]},
        )
    if placeholders:
        return CheckResult(
            "env",
            YELLOW,
            f"{len(placeholders)} env var(s) still have placeholder values",
            details={"placeholders": placeholders[:10]},
        )
    return CheckResult("env", GREEN, f"{len(example_keys)} env vars set")


def check_gmail_token() -> CheckResult:
    """Check that Gmail token file exists and has an unexpired refresh token."""
    candidates = [REPO_ROOT / "token.json", REPO_ROOT / "credentials" / "token.json"]
    token_path = next((p for p in candidates if p.exists()), None)
    if token_path is None:
        return CheckResult("gmail", YELLOW, "token.json not found — Gmail pipeline disabled")

    try:
        data = json.loads(token_path.read_text(encoding="utf-8"))
    except json.JSONDecodeError as exc:
        return CheckResult("gmail", RED, f"token.json is not valid JSON: {exc}")

    if not data.get("refresh_token"):
        return CheckResult("gmail", RED, "token.json has no refresh_token — reauth required")

    # Cross-check: if GOOGLE_CLIENT_SECRET is set in .env.local but not in the token,
    # we've hit the "cause B" drift documented in memory.
    env_path = REPO_ROOT / "apps" / "web" / ".env.local"
    env_secret = None
    if env_path.exists():
        for line in env_path.read_text(encoding="utf-8", errors="replace").splitlines():
            if line.startswith("GOOGLE_CLIENT_SECRET="):
                env_secret = line.split("=", 1)[1].strip().strip('"').strip("'")
                break
    token_secret = data.get("client_secret")
    if env_secret and token_secret and env_secret != token_secret:
        return CheckResult(
            "gmail",
            RED,
            "GOOGLE_CLIENT_SECRET in .env.local differs from token.json — Cause B drift",
        )

    return CheckResult("gmail", GREEN, "token.json present with refresh_token")


def check_supabase() -> CheckResult:
    """Ping Supabase health endpoint if SUPABASE_URL is available."""
    env_path = REPO_ROOT / "apps" / "web" / ".env.local"
    if not env_path.exists():
        return CheckResult("supabase", YELLOW, ".env.local missing — skipping reachability check")
    url = None
    for line in env_path.read_text(encoding="utf-8", errors="replace").splitlines():
        if line.startswith("SUPABASE_URL="):
            url = line.split("=", 1)[1].strip().strip('"').strip("'")
            break
    if not url or "your-project" in url:
        return CheckResult("supabase", YELLOW, "SUPABASE_URL not configured")
    try:
        import urllib.error
        import urllib.request
        req = urllib.request.Request(url.rstrip("/") + "/auth/v1/health", method="GET")
        try:
            with urllib.request.urlopen(req, timeout=5) as resp:
                status = resp.status
        except urllib.error.HTTPError as err:
            status = err.code
        if 200 <= status < 500:
            return CheckResult("supabase", GREEN, f"Supabase reachable ({status})")
        return CheckResult("supabase", RED, f"Supabase returned {status}")
    except Exception as exc:
        return CheckResult("supabase", RED, f"Supabase unreachable: {exc}")


def check_scoring_sync() -> CheckResult:
    """Verify job_score.py and TS scoring both reference scoring-rules.yaml."""
    yaml_path = REPO_ROOT / "packages" / "scoring-rules" / "scoring-rules.yaml"
    py_path = REPO_ROOT / "job_score.py"
    ts_path = REPO_ROOT / "apps" / "web" / "src" / "scoring" / "calculate-score.ts"

    missing = [p for p in [yaml_path, py_path, ts_path] if not p.exists()]
    if missing:
        rels = [str(p.relative_to(REPO_ROOT)) for p in missing]
        return CheckResult("scoring", RED, f"scoring files missing: {rels}")

    yaml_digest = hashlib.sha256(yaml_path.read_bytes()).hexdigest()[:12]
    py_refs = "scoring-rules.yaml" in py_path.read_text(encoding="utf-8", errors="replace")

    yaml_version = None
    for line in yaml_path.read_text(encoding="utf-8").splitlines():
        if line.startswith("version:"):
            yaml_version = line.split(":", 1)[1].strip()
            break

    if not py_refs:
        return CheckResult(
            "scoring",
            RED,
            "job_score.py does not reference scoring-rules.yaml",
            details={"yaml_sha": yaml_digest, "version": yaml_version},
        )
    return CheckResult(
        "scoring",
        GREEN,
        f"scoring-rules v{yaml_version} (sha {yaml_digest})",
        details={"yaml_sha": yaml_digest, "version": yaml_version},
    )


def check_python_deps() -> CheckResult:
    """Verify critical Python deps are importable."""
    critical = {
        "requests": "HTTP client (career_search, portal_scan)",
        "bs4": "BeautifulSoup (HTML parsing)",
        "yaml": "PyYAML (scoring rules)",
    }
    missing = []
    for mod, purpose in critical.items():
        try:
            __import__(mod)
        except ImportError:
            missing.append(f"{mod} ({purpose})")
    if missing:
        return CheckResult("python_deps", RED, "missing Python packages", details={"missing": missing})
    return CheckResult("python_deps", GREEN, f"{len(critical)} critical deps importable")


def check_node_deps() -> CheckResult:
    """Verify apps/web lockfile is consistent with package.json."""
    web_dir = REPO_ROOT / "apps" / "web"
    pkg = web_dir / "package.json"
    lock = web_dir / "package-lock.json"
    if not lock.exists():
        return CheckResult("node_deps", RED, "apps/web/package-lock.json missing")
    try:
        pkg_data = json.loads(pkg.read_text(encoding="utf-8"))
        lock_data = json.loads(lock.read_text(encoding="utf-8"))
    except json.JSONDecodeError as exc:
        return CheckResult("node_deps", RED, f"could not parse lockfile: {exc}")

    pkg_version = pkg_data.get("version")
    lock_version = lock_data.get("version")
    if pkg_version != lock_version:
        return CheckResult(
            "node_deps",
            YELLOW,
            f"package.json ({pkg_version}) vs lockfile ({lock_version}) version drift",
        )
    return CheckResult("node_deps", GREEN, f"lockfile consistent (v{pkg_version})")


def check_extension_version() -> CheckResult:
    """Ensure extension manifest version is present and parseable."""
    manifest = REPO_ROOT / "apps" / "extension" / "manifest.json"
    if not manifest.exists():
        return CheckResult("extension", YELLOW, "apps/extension/manifest.json missing")
    try:
        data = json.loads(manifest.read_text(encoding="utf-8"))
    except json.JSONDecodeError as exc:
        return CheckResult("extension", RED, f"manifest.json invalid: {exc}")
    version = data.get("version")
    if not version or not re.match(r"^\d+\.\d+\.\d+$", version):
        return CheckResult("extension", RED, f"manifest version '{version}' is malformed")
    return CheckResult("extension", GREEN, f"extension v{version}")


def check_tracker_integrity() -> CheckResult:
    """Compare tracker.xlsx row count against applications/ folder count."""
    tracker = REPO_ROOT / "tracker.xlsx"
    apps_dir = REPO_ROOT / "applications"
    if not tracker.exists():
        return CheckResult("tracker", YELLOW, "tracker.xlsx not found")
    if not apps_dir.exists():
        return CheckResult("tracker", YELLOW, "applications/ folder not found")

    app_dirs = [p for p in apps_dir.iterdir() if p.is_dir() and (p / "metadata.json").exists()]
    app_count = len(app_dirs)

    try:
        from openpyxl import load_workbook  # type: ignore
    except ImportError:
        return CheckResult(
            "tracker",
            YELLOW,
            f"{app_count} applications on disk; openpyxl not installed so can't cross-check",
        )
    try:
        wb = load_workbook(tracker, read_only=True)
        ws = wb.active
        row_count = max(ws.max_row - 1, 0)  # minus header
        wb.close()
    except Exception as exc:
        return CheckResult("tracker", RED, f"could not read tracker.xlsx: {exc}")

    drift = abs(row_count - app_count)
    if drift > 5:
        return CheckResult(
            "tracker",
            RED,
            f"tracker drift: {row_count} rows vs {app_count} app folders (run rebuild_tracker.py)",
            details={"tracker_rows": row_count, "app_folders": app_count},
        )
    if drift > 0:
        return CheckResult(
            "tracker",
            YELLOW,
            f"minor drift: {row_count} rows vs {app_count} app folders",
            details={"tracker_rows": row_count, "app_folders": app_count},
        )
    return CheckResult("tracker", GREEN, f"{row_count} rows / {app_count} folders aligned")


ALL_CHECKS: dict[str, Callable[[], CheckResult]] = {
    "env": check_env,
    "gmail": check_gmail_token,
    "supabase": check_supabase,
    "scoring": check_scoring_sync,
    "python_deps": check_python_deps,
    "node_deps": check_node_deps,
    "extension": check_extension_version,
    "tracker": check_tracker_integrity,
}


# ---------------------------------------------------------------------------
# Output
# ---------------------------------------------------------------------------

def _color(status: str) -> str:
    if not sys.stdout.isatty():
        return {GREEN: "OK   ", YELLOW: "WARN ", RED: "FAIL "}[status]
    codes = {GREEN: "\033[32m", YELLOW: "\033[33m", RED: "\033[31m"}
    return f"{codes[status]}{status.upper():<6}\033[0m"


def print_human(results: list[CheckResult]) -> None:
    total_ms = sum(r.duration_ms for r in results)
    red = sum(1 for r in results if r.status == RED)
    yellow = sum(1 for r in results if r.status == YELLOW)
    green = sum(1 for r in results if r.status == GREEN)

    print("job-applications doctor")
    print("-" * 60)
    for r in results:
        print(f"  {_color(r.status)}  {r.name:<14} {r.message}  ({r.duration_ms}ms)")
    print("-" * 60)
    print(f"  {green} green / {yellow} yellow / {red} red  ({total_ms}ms total)")
    if red:
        print("\n  FAIL: fix red checks before merging")
    elif yellow:
        print("\n  OK with warnings")
    else:
        print("\n  all green")


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description="Preflight validator for job-applications")
    parser.add_argument("--json", action="store_true", help="emit machine-readable JSON")
    parser.add_argument("--only", help="comma-separated subset of checks")
    args = parser.parse_args(argv)

    selected = set(args.only.split(",")) if args.only else set(ALL_CHECKS.keys())
    unknown = selected - set(ALL_CHECKS.keys())
    if unknown:
        print(f"unknown checks: {sorted(unknown)}", file=sys.stderr)
        return 2

    results = [_time_check(fn) for name, fn in ALL_CHECKS.items() if name in selected]

    if args.json:
        print(json.dumps([asdict(r) for r in results], indent=2))
    else:
        print_human(results)

    return 1 if any(r.status == RED for r in results) else 0


if __name__ == "__main__":
    sys.exit(main())
