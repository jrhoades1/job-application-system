"""
Rebuild tracker.xlsx from metadata.json files.

The tracker is a derived view of the per-application metadata files. Run this
whenever the tracker is out of sync with applications/ — it rebuilds every
row from scratch in sorted folder order (matching the pairing expected by
tests/test_tracker_sync.py::TestTrackerAccuracy).

Usage: python rebuild_tracker.py
"""
import json
import os
import sys

from openpyxl import Workbook, load_workbook

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
TRACKER_PATH = os.path.join(SCRIPT_DIR, "tracker.xlsx")
APPLICATIONS_DIR = os.path.join(SCRIPT_DIR, "applications")

HEADERS = [
    "Date Applied", "Company", "Role", "Source", "Status",
    "Match Score", "Follow-up Date", "Contact", "Resume Version",
    "Cover Letter", "Notes",
]


def format_status(status):
    if not status:
        return ""
    return status.replace("_", " ").title()


def format_score(match_score):
    if isinstance(match_score, dict):
        overall = match_score.get("overall", "")
    else:
        overall = str(match_score) if match_score else ""
    return overall.replace("_", " ").title() if overall else ""


def preserve_formatting(src_path):
    """If the tracker exists, keep its column widths / styles by loading and
    clearing data rows rather than creating a blank workbook."""
    if not os.path.exists(src_path):
        wb = Workbook()
        ws = wb.active
        ws.title = "Applications"
        for col, header in enumerate(HEADERS, 1):
            ws.cell(row=1, column=col, value=header)
        return wb, ws

    wb = load_workbook(src_path)
    ws = wb["Applications"] if "Applications" in wb.sheetnames else wb.active
    # Clear all data rows (keep row 1 header). Delete from end to avoid reindex.
    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)
    return wb, ws


def main():
    metadata_apps = []
    for folder in sorted(os.listdir(APPLICATIONS_DIR)):
        meta_path = os.path.join(APPLICATIONS_DIR, folder, "metadata.json")
        if os.path.exists(meta_path):
            with open(meta_path, encoding="utf-8") as f:
                metadata_apps.append((folder, json.load(f)))

    wb, ws = preserve_formatting(TRACKER_PATH)

    for i, (folder, meta) in enumerate(metadata_apps, start=2):
        ws.cell(row=i, column=1, value=meta.get("applied_date") or "")
        ws.cell(row=i, column=2, value=meta.get("company", ""))
        ws.cell(row=i, column=3, value=meta.get("role", ""))
        ws.cell(row=i, column=4, value=meta.get("source", "") or "")
        ws.cell(row=i, column=5, value=format_status(meta.get("status", "")))
        ws.cell(row=i, column=6, value=format_score(meta.get("match_score", {})))
        ws.cell(row=i, column=7, value=meta.get("follow_up_date") or "")
        ws.cell(row=i, column=8, value=meta.get("contact", "") or "")
        ws.cell(row=i, column=9, value=meta.get("resume_version") or "")
        ws.cell(row=i, column=10, value=meta.get("cover_letter") or "")
        ws.cell(row=i, column=11, value=meta.get("notes", "") or "")

    wb.save(TRACKER_PATH)
    print(f"Rebuilt tracker with {len(metadata_apps)} rows.")


if __name__ == "__main__":
    sys.exit(main())
