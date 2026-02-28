"""
Test that tracker.xlsx contains all applications from metadata.json files
and that the data is consistent between them.
"""
import os
import json
import pytest
from openpyxl import load_workbook

PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
TRACKER_PATH = os.path.join(PROJECT_ROOT, "tracker.xlsx")
APPLICATIONS_DIR = os.path.join(PROJECT_ROOT, "applications")

# Column mapping (1-indexed)
COL = {
    "applied_date": 1,
    "company": 2,
    "role": 3,
    "source": 4,
    "status": 5,
    "match_score": 6,
    "follow_up_date": 7,
    "contact": 8,
    "resume_version": 9,
    "cover_letter": 10,
    "notes": 11,
}


def load_all_metadata():
    """Load all metadata.json files from application folders."""
    apps = []
    for folder in sorted(os.listdir(APPLICATIONS_DIR)):
        meta_path = os.path.join(APPLICATIONS_DIR, folder, "metadata.json")
        if os.path.exists(meta_path):
            with open(meta_path, encoding="utf-8") as f:
                apps.append((folder, json.load(f)))
    return apps


def load_tracker_rows():
    """Load all data rows from tracker.xlsx as a list of dicts."""
    wb = load_workbook(TRACKER_PATH)
    ws = wb["Applications"]
    rows = []
    for row_num in range(2, ws.max_row + 1):
        company = ws.cell(row=row_num, column=COL["company"]).value
        if not company:
            continue
        rows.append({
            "applied_date": ws.cell(row=row_num, column=COL["applied_date"]).value or "",
            "company": company,
            "role": ws.cell(row=row_num, column=COL["role"]).value or "",
            "source": ws.cell(row=row_num, column=COL["source"]).value or "",
            "status": ws.cell(row=row_num, column=COL["status"]).value or "",
            "match_score": ws.cell(row=row_num, column=COL["match_score"]).value or "",
            "follow_up_date": ws.cell(row=row_num, column=COL["follow_up_date"]).value or "",
            "contact": ws.cell(row=row_num, column=COL["contact"]).value or "",
            "resume_version": ws.cell(row=row_num, column=COL["resume_version"]).value or "",
            "cover_letter": ws.cell(row=row_num, column=COL["cover_letter"]).value or "",
            "notes": ws.cell(row=row_num, column=COL["notes"]).value or "",
        })
    return rows


def format_status(status):
    """Convert snake_case status to Title Case for spreadsheet comparison."""
    if not status:
        return ""
    return status.replace("_", " ").title()


def format_score(match_score):
    """Extract and format the overall match score."""
    if isinstance(match_score, dict):
        overall = match_score.get("overall", "")
    else:
        overall = str(match_score) if match_score else ""
    return overall.replace("_", " ").title() if overall else ""


class TestTrackerExists:
    def test_tracker_file_exists(self):
        assert os.path.exists(TRACKER_PATH), f"tracker.xlsx not found at {TRACKER_PATH}"

    def test_applications_dir_exists(self):
        assert os.path.isdir(APPLICATIONS_DIR), f"applications/ dir not found at {APPLICATIONS_DIR}"

    def test_tracker_has_correct_headers(self):
        wb = load_workbook(TRACKER_PATH)
        ws = wb["Applications"]
        expected_headers = [
            "Date Applied", "Company", "Role", "Source", "Status",
            "Match Score", "Follow-up Date", "Contact", "Resume Version",
            "Cover Letter", "Notes",
        ]
        actual_headers = [ws.cell(row=1, column=i).value for i in range(1, 12)]
        assert actual_headers == expected_headers


class TestNoDuplicateApplications:
    """Detect duplicate company+role entries across application folders.

    Per job-intake Step 0, duplicates should be caught before folder creation.
    If any slip through, this test catches them. Legitimate re-applications
    (after rejection/withdrawal) are allowed — same company+role at different
    statuses where one is rejected/withdrawn is fine.
    """

    def test_no_duplicate_active_applications(self):
        """No two active folders should have the same company+role."""
        metadata_apps = load_all_metadata()
        terminal_statuses = {"rejected", "withdrawn"}

        seen = {}  # (company, role) -> (folder, status)
        duplicates = []

        for folder, meta in metadata_apps:
            key = (meta["company"].lower().strip(), meta["role"].lower().strip())
            status = meta.get("status", "")

            if key in seen:
                prev_folder, prev_status = seen[key]
                # Allow if one of them is terminal (re-application after rejection)
                if prev_status in terminal_statuses or status in terminal_statuses:
                    continue
                duplicates.append(
                    f"  '{meta['company']}' / '{meta['role']}' exists in:\n"
                    f"    - {prev_folder} (status: {prev_status})\n"
                    f"    - {folder} (status: {status})"
                )
            else:
                seen[key] = (folder, status)

        assert not duplicates, (
            "Duplicate active applications found (same company+role, neither rejected/withdrawn):\n"
            + "\n".join(duplicates)
            + "\n\nRun job-intake Step 0 duplicate check to resolve."
        )


class TestTrackerCompleteness:
    def test_every_metadata_has_a_tracker_row(self):
        """Every application folder with metadata.json must have a row in tracker.xlsx."""
        metadata_apps = load_all_metadata()
        tracker_rows = load_tracker_rows()

        tracker_keys = {
            (r["company"].lower().strip(), r["role"].lower().strip())
            for r in tracker_rows
        }

        missing = []
        for folder, meta in metadata_apps:
            key = (meta["company"].lower().strip(), meta["role"].lower().strip())
            if key not in tracker_keys:
                missing.append(f"{meta['company']} - {meta['role']} (folder: {folder})")

        assert not missing, (
            "These applications have metadata.json but are missing from tracker.xlsx:\n"
            + "\n".join(f"  - {m}" for m in missing)
        )

    def test_tracker_row_count_matches_metadata_count(self):
        """Tracker should have exactly as many data rows as there are metadata.json files."""
        metadata_apps = load_all_metadata()
        tracker_rows = load_tracker_rows()
        assert len(tracker_rows) == len(metadata_apps), (
            f"Tracker has {len(tracker_rows)} rows but found {len(metadata_apps)} metadata.json files"
        )

    def test_no_empty_company_rows(self):
        """No tracker row should have an empty company name."""
        tracker_rows = load_tracker_rows()
        for i, row in enumerate(tracker_rows):
            assert row["company"], f"Row {i + 2} has an empty company name"


class TestTrackerAccuracy:
    """Verify that each tracker row matches its corresponding metadata.json.

    The tracker is rebuilt by scanning folders in sorted order, so row N in the
    tracker (after the header) corresponds to the Nth metadata.json in sorted
    folder order. We match by positional index to handle duplicate company+role
    pairs (e.g. an eval copy alongside the real application).
    """

    @pytest.fixture
    def paired_data(self):
        """Pair each metadata.json with its corresponding tracker row by index."""
        metadata_apps = load_all_metadata()
        tracker_rows = load_tracker_rows()
        assert len(metadata_apps) == len(tracker_rows), (
            f"Cannot pair: {len(metadata_apps)} metadata files vs {len(tracker_rows)} tracker rows"
        )
        return list(zip(metadata_apps, tracker_rows))

    def test_company_matches(self, paired_data):
        for (folder, meta), row in paired_data:
            assert row["company"] == meta["company"], (
                f"Folder {folder}: tracker company '{row['company']}' != '{meta['company']}'"
            )

    def test_role_matches(self, paired_data):
        for (folder, meta), row in paired_data:
            assert row["role"] == meta["role"], (
                f"Folder {folder}: tracker role '{row['role']}' != '{meta['role']}'"
            )

    def test_status_matches(self, paired_data):
        for (folder, meta), row in paired_data:
            expected = format_status(meta.get("status", ""))
            assert row["status"] == expected, (
                f"Folder {folder}: tracker status '{row['status']}' != metadata status '{expected}'"
            )

    def test_match_score_matches(self, paired_data):
        for (folder, meta), row in paired_data:
            expected = format_score(meta.get("match_score", {}))
            assert row["match_score"] == expected, (
                f"Folder {folder}: tracker score '{row['match_score']}' != metadata score '{expected}'"
            )

    def test_source_matches(self, paired_data):
        for (folder, meta), row in paired_data:
            expected = meta.get("source", "") or ""
            assert row["source"] == expected, (
                f"Folder {folder}: tracker source '{row['source']}' != metadata source '{expected}'"
            )

    def test_applied_date_matches(self, paired_data):
        for (folder, meta), row in paired_data:
            expected = meta.get("applied_date") or ""
            assert str(row["applied_date"]) == str(expected), (
                f"Folder {folder}: tracker date '{row['applied_date']}' != metadata date '{expected}'"
            )

    def test_resume_version_matches(self, paired_data):
        for (folder, meta), row in paired_data:
            expected = meta.get("resume_version") or ""
            assert str(row["resume_version"]) == str(expected), (
                f"Folder {folder}: tracker resume '{row['resume_version']}' != '{expected}'"
            )

    def test_cover_letter_matches(self, paired_data):
        for (folder, meta), row in paired_data:
            expected = meta.get("cover_letter") or ""
            assert str(row["cover_letter"]) == str(expected), (
                f"Folder {folder}: tracker cover_letter '{row['cover_letter']}' != '{expected}'"
            )
