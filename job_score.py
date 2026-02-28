"""
Email Pipeline — Step 4: Score & Rank

Scores each sourced job description against master/achievements.md
using the same criteria as the job-intake skill, then produces a
ranked review queue.

Creates application folder stubs (status=pending_review) and updates
index.json and tracker.xlsx.

Usage:
    python job_score.py [--rescore]
"""

import json
import os
import re
from datetime import datetime

# Paths
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
PIPELINE_DIR = os.path.join(SCRIPT_DIR, "pipeline")
STAGING_SOURCED = os.path.join(PIPELINE_DIR, "staging", "sourced")
APPLICATIONS_DIR = os.path.join(SCRIPT_DIR, "applications")
INDEX_PATH = os.path.join(APPLICATIONS_DIR, "index.json")
TRACKER_PATH = os.path.join(SCRIPT_DIR, "tracker.xlsx")
ACHIEVEMENTS_PATH = os.path.join(SCRIPT_DIR, "master", "achievements.md")
CONFIG_PATH = os.path.join(SCRIPT_DIR, "pipeline_config.json")
REVIEW_QUEUE_PATH = os.path.join(PIPELINE_DIR, "review_queue.json")
REVIEW_QUEUE_MD = os.path.join(PIPELINE_DIR, "review_queue.md")


def load_config():
    """Load pipeline configuration."""
    if not os.path.exists(CONFIG_PATH):
        return {"auto_skip_rules": {}, "user_preferences": {}}
    with open(CONFIG_PATH, encoding="utf-8") as f:
        return json.load(f)


# ---------------------------------------------------------------------------
# Achievements loading
# ---------------------------------------------------------------------------

def load_achievements(achievements_path=ACHIEVEMENTS_PATH):
    """Parse achievements.md into structured categories.

    Returns {category: [achievement_strings]}.
    """
    if not os.path.exists(achievements_path):
        print(f"  WARNING: achievements.md not found at {achievements_path}")
        return {}

    with open(achievements_path, encoding="utf-8") as f:
        content = f.read()

    achievements = {}
    current_category = None

    for line in content.split("\n"):
        # Category headers: ## Category Name
        cat_match = re.match(r'^##\s+(.+)$', line)
        if cat_match:
            current_category = cat_match.group(1).strip()
            achievements[current_category] = []
            continue

        # Achievement bullet: - text [learned: YYYY-MM-DD]
        bullet_match = re.match(r'^[-*]\s+(.+)$', line)
        if bullet_match and current_category:
            achievement = bullet_match.group(1).strip()
            # Remove the learned tag for matching purposes
            clean = re.sub(r'\s*\[learned:\s*\d{4}-\d{2}-\d{2}\]', '', achievement)
            achievements[current_category].append(clean)

    return achievements


# ---------------------------------------------------------------------------
# Requirements extraction
# ---------------------------------------------------------------------------

def _strip_application_form(text):
    """Remove application form content and EEO boilerplate from description text.

    Many ATS pages include the full application form (dropdowns, checkboxes,
    EEO/disability disclosures) after the actual JD. This pollutes requirement
    extraction with form field options like '-- No answer --'.
    """
    # Cut at common form/apply boundaries
    cutoff_patterns = [
        r'^Apply\s+(?:for this position|Now|Today)',
        r'^Submit\s+Application',
        r'^(?:Required|Optional)\s*\*',
        r'^\*\s*First Name',
        r'^First Name\s*$',
        r'^Human Check',
        r'^Voluntary Self-Identification',
        r'^Invitation for Job Applicants to Self-Identify',
        r'^PUBLIC BURDEN STATEMENT',
        r'^The following questions are entirely optional',
    ]

    lines = text.split('\n')
    cutoff_idx = len(lines)
    for i, line in enumerate(lines):
        stripped = line.strip()
        for pattern in cutoff_patterns:
            if re.match(pattern, stripped, re.IGNORECASE):
                cutoff_idx = i
                break
        if cutoff_idx < len(lines):
            break

    cleaned = '\n'.join(lines[:cutoff_idx])

    # Also strip common EEO/compliance boilerplate paragraphs
    eeo_patterns = [
        r'(?:is an? )?Equal (?:Employment )?Opportunity (?:Employer|and Affirmative Action).*?(?:\n\n|\Z)',
        r'Unfortunately,.*?(?:not currently hiring|Territories)\.',
    ]
    for pattern in eeo_patterns:
        cleaned = re.sub(pattern, '', cleaned, flags=re.IGNORECASE | re.DOTALL)

    return cleaned.strip()


def extract_requirements(job_description):
    """Extract structured requirements from job description text.

    Returns {hard_requirements, preferred, responsibilities, keywords, red_flags}.
    """
    if not job_description:
        return {"hard_requirements": [], "preferred": [], "responsibilities": [],
                "keywords": [], "red_flags": []}

    # Strip application form and EEO boilerplate before parsing
    cleaned_description = _strip_application_form(job_description)
    lines = cleaned_description.split("\n")

    hard_requirements = []
    preferred = []
    responsibilities = []
    keywords = set()

    current_section = None
    section_patterns = {
        "requirements": r"(?:requirements?|qualifications?|what you.?(?:ll)?\s*need|must have|minimum|experience|education|skills?\s+(?:and|&)\s+(?:knowledge|skills)|specialized knowledge|technical skills|what (?:we|you).+(?:look|need)|who you are)",
        "preferred": r"(?:preferred|nice to have|bonus|desired|plus|ideally|good to have|additional|differenti)",
        "responsibilities": r"(?:responsibilities|what you.?(?:ll)?\s*do|duties|role|about the (?:role|position)|key (?:areas|functions)|you will|your (?:impact|mission|role))",
    }

    for line in lines:
        stripped = line.strip()
        if not stripped:
            continue

        # Detect section headers
        lower = stripped.lower()
        for section, pattern in section_patterns.items():
            if re.search(pattern, lower) and len(stripped) < 80:
                current_section = section
                break

        # Extract bullet points
        bullet_match = re.match(r'^[-•*]\s*(.+)$', stripped)
        if not bullet_match:
            # Also match numbered items
            bullet_match = re.match(r'^\d+[.)]\s*(.+)$', stripped)

        if bullet_match:
            item = bullet_match.group(1).strip()

            if current_section == "requirements":
                hard_requirements.append(item)
            elif current_section == "preferred":
                preferred.append(item)
            elif current_section == "responsibilities":
                responsibilities.append(item)
            else:
                # Guess based on content
                if _is_requirement(item):
                    hard_requirements.append(item)
                elif _is_preferred(item):
                    preferred.append(item)
                else:
                    responsibilities.append(item)
        elif current_section == "requirements" and len(stripped) > 15 and len(stripped) < 200:
            # Non-bullet lines in a requirements section (plain text requirements)
            if not re.match(r'^(?:Travel|Note|Image|About|Share)\b', stripped):
                hard_requirements.append(stripped)
        elif current_section == "preferred" and len(stripped) > 15 and len(stripped) < 200:
            if not re.match(r'^(?:Travel|Note|Image|About|Share)\b', stripped):
                preferred.append(stripped)

    # Extract keywords (use original for broader keyword detection)
    keywords = _extract_keywords(job_description)

    # Detect red flags (use cleaned to avoid false positives from form content)
    red_flags = _detect_red_flags(cleaned_description)

    return {
        "hard_requirements": hard_requirements,
        "preferred": preferred,
        "responsibilities": responsibilities,
        "keywords": list(keywords),
        "red_flags": red_flags,
    }


def _is_requirement(text):
    """Check if a bullet point is likely a requirement."""
    # Skip obvious non-requirements
    skip_indicators = [
        r'(?:no special physical demands|travel up to|office environment)',
        r'(?:-- No answer|background check|drug screen|e-verify)',
        r'(?:salary|compensation|benefits|401|pto|paid time)',
        r'(?:equal (?:employment )?opportunity|affirmative action)',
        r'(?:visa sponsorship|legally eligible)',
    ]
    if any(re.search(p, text, re.IGNORECASE) for p in skip_indicators):
        return False

    req_indicators = [
        r'\d+\+?\s*years?', r'(?:must|required|minimum)', r'(?:degree|bachelor|master|phd)',
        r'(?:experience (?:with|in|leading|building|managing|developing|driving))',
        r'(?:proficiency in|expertise in|deep expertise|proven)',
        r'(?:certification|certified)',
        r'(?:track record|demonstrated|strong (?:strategic|technical|communication))',
        r'(?:knowledge of|ability to|skilled in|familiarity with)',
    ]
    return any(re.search(p, text, re.IGNORECASE) for p in req_indicators)


def _is_preferred(text):
    """Check if a bullet point is likely a preferred qualification."""
    pref_indicators = [
        r'(?:preferred|nice to have|plus|bonus|ideally|advantageous)',
        r'(?:familiarity with|exposure to|knowledge of)',
    ]
    return any(re.search(p, text, re.IGNORECASE) for p in pref_indicators)


def _extract_keywords(text):
    """Extract technology and skill keywords from job description."""
    keyword_patterns = [
        # Technologies
        r'\b(?:Python|Java|JavaScript|TypeScript|Go|Rust|C\+\+|Ruby|Scala|Kotlin)\b',
        r'\b(?:AWS|Azure|GCP|Google Cloud|Kubernetes|Docker|Terraform)\b',
        r'\b(?:React|Angular|Vue|Next\.js|Node\.js|FastAPI|Django|Flask|Spring)\b',
        r'\b(?:PostgreSQL|MySQL|MongoDB|Redis|DynamoDB|Elasticsearch)\b',
        r'\b(?:CI/CD|DevOps|Agile|Scrum|Kanban)\b',
        # Healthcare specific
        r'\b(?:HL7|FHIR|DICOM|HIPAA|SOC2|HITRUST|EHR|EMR)\b',
        r'\b(?:PHI|PII|FDA|CMS|ICD-10)\b',
        # AI/ML
        r'\b(?:AI|ML|NLP|LLM|GPT|machine learning|deep learning|neural network)\b',
        r'\b(?:TensorFlow|PyTorch|scikit-learn|LangChain)\b',
        # Leadership
        r'\b(?:microservices|scalability|architecture|system design)\b',
        r'\b(?:team building|mentoring|roadmap|OKR|KPI)\b',
    ]

    keywords = set()
    for pattern in keyword_patterns:
        matches = re.findall(pattern, text, re.IGNORECASE)
        keywords.update(m.strip() for m in matches)

    return keywords


def _detect_red_flags(text):
    """Detect red flags in the job description."""
    flags = []
    lower = text.lower()

    red_flag_patterns = [
        (r"wear many hats", "Vague role scope — 'wear many hats'"),
        (r"fast[- ]paced", "Fast-paced environment (potential burnout signal)"),
        (r"must be willing to work (?:nights|weekends|overtime)", "Expects overtime"),
        (r"(?:ninja|rockstar|guru|wizard|unicorn)", "Buzzword-heavy role description"),
        (r"unlimited (?:pto|vacation)", "Unlimited PTO (often means less PTO taken)"),
        (r"competitive salary", "No salary range listed — 'competitive salary'"),
    ]

    for pattern, flag in red_flag_patterns:
        if re.search(pattern, lower):
            flags.append(flag)

    # Check for unrealistic experience combos
    years_matches = re.findall(r'(\d+)\+?\s*years?', lower)
    if years_matches:
        max_years = max(int(y) for y in years_matches)
        if max_years > 15:
            flags.append(f"Requires {max_years}+ years — unusually high")

    return flags


# ---------------------------------------------------------------------------
# Scoring
# ---------------------------------------------------------------------------

def _simple_stem(word):
    """Reduce word to approximate stem for matching."""
    if len(word) <= 3:
        return word
    if word.endswith('ing') and len(word) > 5:
        return word[:-3]
    if word.endswith('tion') and len(word) > 6:
        return word[:-4]
    if word.endswith('ed') and len(word) > 4:
        return word[:-2]
    if word.endswith('ment') and len(word) > 6:
        return word[:-4]
    if word.endswith('ness') and len(word) > 6:
        return word[:-4]
    if word.endswith('ies') and len(word) > 4:
        return word[:-3] + 'y'
    if word.endswith('s') and not word.endswith('ss') and len(word) > 4:
        return word[:-1]
    return word


def score_requirement(requirement, achievements):
    """Score a single requirement against the achievements inventory.

    Returns {requirement, match_type: 'strong'|'partial'|'gap', evidence, category}.
    """
    req_lower = requirement.lower()

    # Extract key terms and stem them for better matching
    req_words = re.findall(r'\b[a-z]{3,}\b', req_lower)
    req_terms = set(_simple_stem(w) for w in req_words)

    best_match = None
    best_score = 0

    for category, items in achievements.items():
        for item in items:
            item_lower = item.lower()
            item_words = re.findall(r'\b[a-z]{3,}\b', item_lower)
            item_terms = set(_simple_stem(w) for w in item_words)

            # Calculate overlap
            if req_terms and item_terms:
                overlap = len(req_terms & item_terms) / max(len(req_terms), 1)
            else:
                overlap = 0

            # Check for direct keyword matches (technologies, frameworks)
            direct_keywords = re.findall(
                r'\b(?:Python|Java|AWS|Azure|GCP|Kubernetes|Docker|Terraform|'
                r'React|Node|HIPAA|SOC2|FHIR|HL7|DICOM|AI|ML|NLP|'
                r'microservices|agile|scrum|DevOps|CI/CD)\b',
                requirement, re.IGNORECASE
            )
            for kw in direct_keywords:
                if kw.lower() in item_lower:
                    overlap += 0.3

            # Check for experience level matches
            years_req = re.search(r'(\d+)\+?\s*years?', req_lower)
            years_ach = re.search(r'(\d+)\+?\s*years?', item_lower)
            if years_req and years_ach:
                if int(years_ach.group(1)) >= int(years_req.group(1)):
                    overlap += 0.2

            if overlap > best_score:
                best_score = overlap
                best_match = {"evidence": item, "category": category}

    if best_score >= 0.35:
        return {
            "requirement": requirement,
            "match_type": "strong",
            "evidence": best_match["evidence"],
            "category": best_match["category"],
        }
    elif best_score >= 0.2:
        return {
            "requirement": requirement,
            "match_type": "partial",
            "evidence": best_match["evidence"] if best_match else "",
            "category": best_match["category"] if best_match else "",
        }
    else:
        return {
            "requirement": requirement,
            "match_type": "gap",
            "evidence": "",
            "category": "",
        }


def calculate_overall_score(matches):
    """Apply the scoring thresholds from packages/scoring-rules/scoring-rules.yaml.

    Thresholds shared with TypeScript web app (apps/web/src/scoring/calculate-score.ts).
    If you change thresholds here, update the YAML and TypeScript to match.

    Strong: 80%+ matched, 0 critical gaps
    Good: 60-80%, 0-1 addressable
    Stretch: 40-60%, 1-2 addressable
    Long shot: below 40%, multiple hard gaps
    """
    if not matches:
        return {
            "overall": "long_shot",
            "match_percentage": 0,
            "strong_count": 0,
            "partial_count": 0,
            "gap_count": 0,
        }

    strong = [m for m in matches if m["match_type"] == "strong"]
    partial = [m for m in matches if m["match_type"] == "partial"]
    gaps = [m for m in matches if m["match_type"] == "gap"]

    total = len(matches)
    match_pct = (len(strong) + len(partial) * 0.5) / total if total > 0 else 0

    if match_pct >= 0.8 and len(gaps) == 0:
        overall = "strong"
    elif match_pct >= 0.6 and len(gaps) <= 1:
        overall = "good"
    elif match_pct >= 0.4 and len(gaps) <= 2:
        overall = "stretch"
    else:
        overall = "long_shot"

    return {
        "overall": overall,
        "match_percentage": round(match_pct * 100, 1),
        "strong_count": len(strong),
        "partial_count": len(partial),
        "gap_count": len(gaps),
    }


def detect_employment_type(description):
    """Detect contract/temp/part-time from description text."""
    if not description:
        return "unknown"

    # Strip EEO/compliance boilerplate to avoid false positives
    # (e.g., "federal contractor" in EEO statements)
    cleaned = _strip_application_form(description)
    lower = cleaned.lower()

    # Check for contract but exclude false positives from compliance context
    contract_match = re.search(r'\b(?:contract|contractor|c2c|w2|1099)\b', lower)
    if contract_match:
        # Check if it's in a compliance/EEO context
        context_start = max(0, contract_match.start() - 80)
        context = lower[context_start:contract_match.end() + 40]
        eeo_context = any(phrase in context for phrase in [
            'federal contract', 'government contract', 'contract compliance',
            'subcontractor', 'affirmative action',
        ])
        if not eeo_context:
            return "contract"

    if re.search(r'\b(?:part[- ]time)\b', lower):
        return "part_time"
    if re.search(r'\b(?:temporary|temp position|temp role)\b', lower):
        return "temp"
    if re.search(r'\b(?:full[- ]time|permanent)\b', lower):
        return "full_time"

    return "full_time"  # Default assumption


def detect_location_match(description, user_preferences):
    """Check if location requirements match user preferences."""
    if not description:
        return {"match": True, "location": "Unknown", "remote_status": "unknown"}

    lower = description.lower()
    preferred_location = (user_preferences.get("location") or "").lower()

    # Detect remote status
    if re.search(r'\b(?:fully remote|100% remote|remote[- ]first|work from (?:home|anywhere))\b', lower):
        remote_status = "remote"
    elif re.search(r'\bhybrid\b', lower):
        remote_status = "hybrid"
    elif re.search(r'\b(?:on[- ]?site|in[- ]?office|in[- ]?person)\b', lower):
        remote_status = "onsite"
    else:
        remote_status = "unknown"

    # Check match
    if "remote" in preferred_location:
        match = remote_status in ("remote", "unknown")
    else:
        match = True  # Can't reliably geo-match, default to pass

    # Extract location text
    location = ""
    loc_match = re.search(r'(?:location|based in|located in)[:\s]+([^\n.]{3,50})', description, re.IGNORECASE)
    if loc_match:
        location = loc_match.group(1).strip()

    return {"match": match, "location": location, "remote_status": remote_status}


# ---------------------------------------------------------------------------
# Auto-skip rules
# ---------------------------------------------------------------------------

def check_auto_skip(lead, score_result, auto_skip_rules, user_preferences, employment_type):
    """Apply auto-skip rules from config.

    Returns skip reason if should be skipped, None otherwise.
    """
    if not auto_skip_rules:
        return None

    # Check minimum score
    min_score = auto_skip_rules.get("min_score")
    if min_score:
        score_order = {"strong": 4, "good": 3, "stretch": 2, "long_shot": 1}
        if score_order.get(score_result["overall"], 0) < score_order.get(min_score, 0):
            return f"Below minimum score threshold ({score_result['overall']} < {min_score})"

    # Check excluded employment types
    excluded_types = auto_skip_rules.get("excluded_employment_types", [])
    if employment_type in excluded_types:
        return f"Employment type: {employment_type} (auto-skip rule)"

    # Check excluded companies
    excluded_companies = [c.lower() for c in auto_skip_rules.get("excluded_companies", [])]
    if lead.get("company", "").lower() in excluded_companies:
        return "Company in exclusion list"

    return None


# ---------------------------------------------------------------------------
# Deduplication
# ---------------------------------------------------------------------------

def load_index():
    """Load the applications index.json."""
    if os.path.exists(INDEX_PATH):
        with open(INDEX_PATH, encoding="utf-8") as f:
            return json.load(f)
    return []


def check_existing_application(company, role, index):
    """Check if this company+role already exists in the system.

    Returns the existing entry dict or None.
    """
    key = (company.lower().strip(), role.lower().strip())

    for entry in index:
        entry_key = (entry.get("company", "").lower().strip(),
                     entry.get("role", "").lower().strip())
        if entry_key == key:
            return entry

    return None


# ---------------------------------------------------------------------------
# Ranking
# ---------------------------------------------------------------------------

def rank_jobs(scored_leads):
    """Rank scored leads for user review.

    Primary: score tier (strong > good > stretch > long_shot)
    Secondary: match percentage (descending)
    Tertiary: gap count (ascending)
    Tiebreaker: company name alphabetical
    """
    score_order = {"strong": 0, "good": 1, "stretch": 2, "long_shot": 3}

    def sort_key(lead):
        score = lead.get("score_result", {})
        return (
            score_order.get(score.get("overall", "long_shot"), 3),
            -score.get("match_percentage", 0),
            score.get("gap_count", 99),
            lead.get("company", "").lower(),
        )

    ranked = sorted(scored_leads, key=sort_key)
    for i, lead in enumerate(ranked):
        lead["rank"] = i + 1

    return ranked


# ---------------------------------------------------------------------------
# Application folder creation
# ---------------------------------------------------------------------------

def slugify(text):
    """Convert text to a URL-friendly slug. Matches swooped_import.py."""
    text = text.lower().strip()
    text = re.sub(r'[^\w\s-]', '', text)
    text = re.sub(r'[\s_]+', '-', text)
    text = re.sub(r'-+', '-', text)
    return text.strip('-')[:60]


def create_pipeline_metadata(lead, score_result, requirements, employment_type, location_info):
    """Create a metadata.json dict for a pipeline-sourced application."""
    today = datetime.now().strftime("%Y-%m-%d")

    return {
        "company": lead.get("company", ""),
        "role": lead.get("role", ""),
        "location": location_info.get("location", ""),
        "compensation": lead.get("compensation"),
        "applied_date": None,
        "source": lead.get("source_platform", "Email Pipeline"),
        "source_url": lead.get("career_page_url", ""),
        "status": "pending_review",
        "follow_up_date": None,
        "contact": "",
        "resume_version": None,
        "cover_letter": None,
        "former_employer": False,
        "former_employer_role": None,
        "notes": f"Sourced via email pipeline on {today}. "
                 f"Email subject: {lead.get('raw_subject', 'N/A')[:80]}",
        "match_score": {
            "overall": score_result.get("overall", ""),
            "requirements_matched": [m["requirement"] for m in lead.get("matches", [])
                                      if m["match_type"] == "strong"],
            "requirements_partial": [m["requirement"] for m in lead.get("matches", [])
                                      if m["match_type"] == "partial"],
            "gaps": [m["requirement"] for m in lead.get("matches", [])
                     if m["match_type"] == "gap"],
            "addressable_gaps": [],
            "hard_gaps": [],
            "keywords": requirements.get("keywords", []),
        },
        "tailoring_intensity": None,
        "interview_date": None,
        "interview_round": None,
        "interview_type": None,
        "interview_notes_file": None,
        "rejection_date": None,
        "rejection_reason": None,
        "rejection_insights": None,
        "offer": {
            "salary": None,
            "equity": None,
            "signing_bonus": None,
            "remote": location_info.get("remote_status"),
            "benefits_notes": None,
            "decision_deadline": None,
        },
        "offer_accepted": None,
        "learning_flags": [],
        # Pipeline-specific fields
        "email_uid": lead.get("email_uid"),
        "pipeline_batch": lead.get("pipeline_batch"),
        "pipeline_confidence": lead.get("confidence"),
        "employment_type": employment_type,
        "skip_date": None,
        "skip_reason": None,
    }


def create_application_stubs(ranked_leads, index):
    """Create application folders for scored leads.

    Returns list of created folder info. Skips duplicates.
    """
    os.makedirs(APPLICATIONS_DIR, exist_ok=True)
    today = datetime.now().strftime("%Y-%m-%d")

    created = []
    for lead in ranked_leads:
        company = lead.get("company", "Unknown")
        role = lead.get("role", "Unknown")

        # Check for existing application
        existing = check_existing_application(company, role, index)
        if existing:
            status = existing.get("status", "")
            if status in ("rejected", "withdrawn"):
                lead["dedup_note"] = f"Previously {status} — consider re-applying?"
            else:
                lead["dedup_note"] = f"Already tracked as '{status}'"
                lead["skipped_dedup"] = True
                continue

        # Create folder
        company_slug = slugify(company)
        role_slug = slugify(role)
        folder_name = f"{today}_{company_slug}_{role_slug}"
        folder_path = os.path.join(APPLICATIONS_DIR, folder_name)

        if os.path.exists(folder_path):
            lead["dedup_note"] = "Folder already exists"
            lead["skipped_dedup"] = True
            continue

        os.makedirs(folder_path, exist_ok=True)

        # Create metadata.json
        metadata = create_pipeline_metadata(
            lead,
            lead.get("score_result", {}),
            lead.get("requirements", {}),
            lead.get("employment_type", "full_time"),
            lead.get("location_info", {}),
        )

        meta_path = os.path.join(folder_path, "metadata.json")
        with open(meta_path, "w", encoding="utf-8") as f:
            json.dump(metadata, f, indent=2, ensure_ascii=False)

        # Save job description
        description = lead.get("description_text", "")
        if description:
            jd_path = os.path.join(folder_path, "job-description.md")
            with open(jd_path, "w", encoding="utf-8") as f:
                f.write(f"# {role} — {company}\n\n")
                f.write(f"**Source:** Email pipeline ({lead.get('source_platform', 'Unknown')})\n")
                f.write(f"**Career Page:** {lead.get('career_page_url', 'N/A')}\n")
                f.write(f"**Scraped:** {today}\n\n---\n\n")
                f.write(description)

        lead["application_folder"] = folder_name
        lead["folder_path"] = folder_path

        # Update index
        index.append({
            "company": company,
            "role": role,
            "status": "pending_review",
            "folder": folder_name,
        })

        created.append({
            "company": company,
            "role": role,
            "folder": folder_name,
            "score": lead.get("score_result", {}).get("overall", ""),
        })

    return created


def update_index_and_tracker(index, new_entries):
    """Save updated index and add new rows to tracker.xlsx."""
    # Save index
    with open(INDEX_PATH, "w", encoding="utf-8") as f:
        json.dump(index, f, indent=2, ensure_ascii=False)

    # Update tracker
    try:
        from openpyxl import load_workbook, Workbook
    except ImportError:
        print("  WARNING: openpyxl not installed, skipping tracker update")
        return

    if os.path.exists(TRACKER_PATH):
        wb = load_workbook(TRACKER_PATH)
        if "Applications" in wb.sheetnames:
            ws = wb["Applications"]
        else:
            ws = wb.active
            ws.title = "Applications"
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Applications"
        headers = ["Date Applied", "Company", "Role", "Source", "Status",
                    "Match Score", "Follow-up Date", "Contact", "Resume Version",
                    "Cover Letter", "Notes"]
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)

    # Build existing lookup
    existing = set()
    for row_num in range(2, ws.max_row + 1):
        company = ws.cell(row=row_num, column=2).value
        role = ws.cell(row=row_num, column=3).value
        if company and role:
            existing.add((company.lower().strip(), role.lower().strip()))

    # Add new rows
    added = 0
    for entry in new_entries:
        key = (entry["company"].lower().strip(), entry["role"].lower().strip())
        if key not in existing:
            new_row = ws.max_row + 1
            ws.cell(row=new_row, column=1, value="")  # No applied date yet
            ws.cell(row=new_row, column=2, value=entry["company"])
            ws.cell(row=new_row, column=3, value=entry["role"])
            ws.cell(row=new_row, column=4, value="Email Pipeline")
            ws.cell(row=new_row, column=5, value="Pending Review")
            ws.cell(row=new_row, column=6, value=entry.get("score", "").replace("_", " ").title())
            ws.cell(row=new_row, column=11, value="Sourced via email pipeline")
            existing.add(key)
            added += 1

    wb.save(TRACKER_PATH)
    print(f"  Tracker updated: {added} new rows")


# ---------------------------------------------------------------------------
# Review queue generation
# ---------------------------------------------------------------------------

def generate_review_queue(ranked_leads, auto_skipped, unresolved):
    """Write review_queue.json and review_queue.md for the email-triage skill."""
    os.makedirs(PIPELINE_DIR, exist_ok=True)

    batch_id = f"{datetime.now().strftime('%Y-%m-%d')}_{os.urandom(3).hex()}"

    queue = {
        "batch_id": batch_id,
        "generated_at": datetime.now().isoformat(),
        "leads": [],
        "auto_skipped": auto_skipped,
        "unresolved": unresolved,
    }

    for lead in ranked_leads:
        score = lead.get("score_result", {})
        queue["leads"].append({
            "rank": lead.get("rank", 0),
            "company": lead.get("company", ""),
            "role": lead.get("role", ""),
            "score": {
                "overall": score.get("overall", ""),
                "match_percentage": score.get("match_percentage", 0),
                "strong_count": score.get("strong_count", 0),
                "partial_count": score.get("partial_count", 0),
                "gap_count": score.get("gap_count", 0),
            },
            "top_matches": [m["requirement"][:60] for m in lead.get("matches", [])
                            if m["match_type"] == "strong"][:3],
            "top_gaps": [m["requirement"][:60] for m in lead.get("matches", [])
                         if m["match_type"] == "gap"][:2],
            "source_platform": lead.get("source_platform", ""),
            "email_uid": lead.get("email_uid", ""),
            "email_date": lead.get("email_date", ""),
            "career_page_url": lead.get("career_page_url", ""),
            "application_folder": lead.get("application_folder", ""),
            "employment_type": lead.get("employment_type", "full_time"),
            "location": lead.get("location_info", {}).get("location", ""),
            "remote_status": lead.get("location_info", {}).get("remote_status", "unknown"),
            "compensation": lead.get("compensation"),
            "confidence": lead.get("confidence", 0),
            "red_flags": lead.get("red_flags", []),
            "dedup_note": lead.get("dedup_note", ""),
            "status": "pending_review",
        })

    # Save JSON
    with open(REVIEW_QUEUE_PATH, "w", encoding="utf-8") as f:
        json.dump(queue, f, indent=2, ensure_ascii=False)

    # Generate markdown summary
    _write_review_md(queue)

    return queue


def _write_review_md(queue):
    """Write a human-readable markdown summary of the review queue."""
    lines = [
        "# Email Pipeline Review Queue",
        "",
        f"**Generated:** {queue['generated_at'][:16]}",
        f"**Batch:** {queue['batch_id']}",
        f"**Leads:** {len(queue['leads'])} | **Auto-skipped:** {len(queue['auto_skipped'])} | "
        f"**Unresolved:** {len(queue['unresolved'])}",
        "",
    ]

    # Group by score tier
    tiers = {"strong": [], "good": [], "stretch": [], "long_shot": []}
    for lead in queue["leads"]:
        tier = lead["score"]["overall"]
        tiers.get(tier, tiers["long_shot"]).append(lead)

    for tier_name, tier_leads in tiers.items():
        if not tier_leads:
            continue

        lines.append(f"## {tier_name.replace('_', ' ').title()} ({len(tier_leads)})")
        lines.append("")

        for lead in tier_leads:
            rank = lead["rank"]
            company = lead["company"]
            role = lead["role"]
            pct = lead["score"]["match_percentage"]
            source = lead["source_platform"]
            flags = []
            if lead.get("red_flags"):
                flags.append("RED FLAGS")
            if lead.get("employment_type") not in ("full_time", "unknown"):
                flags.append(lead["employment_type"].upper())
            if lead.get("dedup_note"):
                flags.append(lead["dedup_note"])

            flag_str = f" [{', '.join(flags)}]" if flags else ""

            lines.append(f"**[{rank}] {company} — {role}**")
            lines.append(f"  Score: {pct}% match | Source: {source}{flag_str}")

            if lead.get("top_matches"):
                lines.append(f"  Matches: {', '.join(lead['top_matches'][:3])}")
            if lead.get("top_gaps"):
                lines.append(f"  Gaps: {', '.join(lead['top_gaps'][:2])}")
            lines.append("")

    # Unresolved
    if queue["unresolved"]:
        lines.append(f"## Unresolved ({len(queue['unresolved'])})")
        lines.append("")
        for item in queue["unresolved"]:
            lines.append(f"- **{item.get('company', 'Unknown')}** — {item.get('role', 'Unknown')}: "
                         f"{item.get('reason', 'Unknown error')}")
        lines.append("")

    # Auto-skipped
    if queue["auto_skipped"]:
        lines.append(f"## Auto-Skipped ({len(queue['auto_skipped'])})")
        lines.append("")
        for item in queue["auto_skipped"]:
            lines.append(f"- {item.get('company', 'Unknown')} — {item.get('role', 'Unknown')}: "
                         f"{item.get('reason', 'Rule match')}")
        lines.append("")

    with open(REVIEW_QUEUE_MD, "w", encoding="utf-8") as f:
        f.write("\n".join(lines))


# ---------------------------------------------------------------------------
# Main pipeline
# ---------------------------------------------------------------------------

def main():
    import argparse

    parser = argparse.ArgumentParser(description="Score and rank sourced job descriptions")
    parser.add_argument("--rescore", action="store_true", help="Re-score already scored leads")
    parser.parse_args()

    print("=" * 60)
    print("  EMAIL PIPELINE — STEP 4: SCORE & RANK")
    print("=" * 60)

    config = load_config()
    auto_skip_rules = config.get("auto_skip_rules", {})
    user_preferences = config.get("user_preferences", {})

    # Load achievements
    print("\n  Loading achievements...")
    achievements = load_achievements()
    if not achievements:
        print("  WARNING: No achievements loaded — scoring will be limited")
    else:
        total_ach = sum(len(v) for v in achievements.values())
        print(f"  Loaded {total_ach} achievements across {len(achievements)} categories")

    # Load sourced results
    if not os.path.exists(STAGING_SOURCED):
        print("  No sourced leads found.")
        return

    sourced_files = [f for f in os.listdir(STAGING_SOURCED) if f.endswith(".json")]
    if not sourced_files:
        print("  No sourced lead files found.")
        return

    print(f"\n  Scoring {len(sourced_files)} sourced leads...")

    # Load index for dedup
    index = load_index()

    scored_leads = []
    auto_skipped = []
    unresolved = []

    batch_id = f"{datetime.now().strftime('%Y-%m-%d')}_{os.urandom(3).hex()}"

    for filename in sorted(sourced_files):
        filepath = os.path.join(STAGING_SOURCED, filename)
        with open(filepath, encoding="utf-8") as f:
            sourced = json.load(f)

        # Skip unresolved leads
        if sourced.get("status") == "unresolved":
            unresolved.append({
                "company": sourced.get("lead", {}).get("company", "Unknown"),
                "role": sourced.get("lead", {}).get("role", "Unknown"),
                "reason": sourced.get("unresolved_reason", "Unknown"),
                "email_uid": sourced.get("lead", {}).get("email_uid", ""),
            })
            continue

        lead = sourced.get("lead", {})
        scraped = sourced.get("scraped", {})
        description = scraped.get("description_text", "")

        if not description:
            unresolved.append({
                "company": lead.get("company", "Unknown"),
                "role": lead.get("role", "Unknown"),
                "reason": "No description text available",
                "email_uid": lead.get("email_uid", ""),
            })
            continue

        # Extract requirements (form content is stripped internally)
        requirements = extract_requirements(description)

        # Score each requirement against achievements
        all_reqs = requirements["hard_requirements"] + requirements["preferred"]
        if not all_reqs:
            # If no structured requirements found, treat cleaned description as context
            cleaned_desc = _strip_application_form(description)
            all_reqs = [line.strip() for line in cleaned_desc.split("\n")
                        if len(line.strip()) > 20 and _is_requirement(line)][:15]

        matches = [score_requirement(req, achievements) for req in all_reqs]
        score_result = calculate_overall_score(matches)

        # Detect employment type
        employment_type = detect_employment_type(description)

        # Check location
        location_info = detect_location_match(description, user_preferences)

        # Build the scored lead
        scored_lead = {
            "company": lead.get("company", ""),
            "role": lead.get("role", ""),
            "source_platform": lead.get("source_platform", ""),
            "email_uid": lead.get("email_uid", ""),
            "email_date": lead.get("email_date", ""),
            "raw_subject": lead.get("raw_subject", ""),
            "confidence": lead.get("confidence", 0),
            "career_page_url": scraped.get("url", ""),
            "description_text": description,
            "compensation": scraped.get("compensation"),
            "score_result": score_result,
            "matches": matches,
            "requirements": requirements,
            "employment_type": employment_type,
            "location_info": location_info,
            "red_flags": requirements.get("red_flags", []),
            "pipeline_batch": batch_id,
        }

        # Check auto-skip rules
        skip_reason = check_auto_skip(scored_lead, score_result, auto_skip_rules,
                                       user_preferences, employment_type)
        if skip_reason:
            auto_skipped.append({
                "company": lead.get("company", ""),
                "role": lead.get("role", ""),
                "reason": skip_reason,
                "score": score_result.get("overall", ""),
                "email_uid": lead.get("email_uid", ""),
            })
            continue

        scored_leads.append(scored_lead)

    # Rank
    print(f"\n  Ranking {len(scored_leads)} scored leads...")
    ranked = rank_jobs(scored_leads)

    # Score distribution
    dist = {}
    for lead in ranked:
        tier = lead.get("score_result", {}).get("overall", "unknown")
        dist[tier] = dist.get(tier, 0) + 1

    print(f"  Score distribution: {json.dumps(dist)}")

    # Create application stubs
    print("\n  Creating application folders...")
    created = create_application_stubs(ranked, index)
    print(f"  Created {len(created)} new application folders")

    # Update index and tracker
    print("\n  Updating index and tracker...")
    update_index_and_tracker(index, created)

    # Generate review queue
    print("\n  Generating review queue...")
    queue = generate_review_queue(ranked, auto_skipped, unresolved)

    # Summary
    print("\n  Results:")
    print(f"    Scored:       {len(scored_leads)}")
    print(f"    Auto-skipped: {len(auto_skipped)}")
    print(f"    Unresolved:   {len(unresolved)}")
    print(f"    Folders made: {len(created)}")
    print(f"    Review queue: {REVIEW_QUEUE_PATH}")
    print(f"    Review MD:    {REVIEW_QUEUE_MD}")

    for tier in ["strong", "good", "stretch", "long_shot"]:
        count = dist.get(tier, 0)
        if count:
            print(f"      {tier}: {count}")

    print(f"\n{'=' * 60}")
    print(f"  SCORING COMPLETE — {len(queue['leads'])} leads ready for review")
    print(f"{'=' * 60}")


if __name__ == "__main__":
    main()
